import csv
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.exceptions import ReadOnlyWorkbookException
# from openpyxl.utils.exceptions import *


def printCell(sheet):
    max_row = sheet.max_row
    max_col = sheet.max_column
    for i in range(1, max_row+1):
        # iterate over all columns
        for j in range(1, max_col+1):
            # get particular cell value
            cell_obj = sheet.cell(row=i, column=j)
            # print cell value
            print(cell_obj.value, end='')
            if j != max_col:
                print(' | ', end='')
        # print new line
        print('\n')


file = r"C:\Users\tsezg523\Desktop\testfile.xlsx"

wbs = load_workbook(file, read_only=False)
ws1 = wbs["Sheet1"]

#sheet = wbs.active
#cell = sheet.cell(row=2, column=1)

print("Sheet names:", wbs.sheetnames)

print("Sheet1:")
printCell(wbs["Sheet1"])
print("Sheet2:")
printCell(wbs["Sheet2"])

print("Change the price for the first product in both Sheet1 and Sheet2:")
try:
    wbs["Sheet1"]['B2'] = 321
    wbs["Sheet2"]['B'][2].value = 876
    wbs["Sheet1"][2][1].value = 111
    printCell(wbs["Sheet1"])
    print("Make the change into the excel file:")
    wbs.save(file)
except ReadOnlyWorkbookException:
    print("The file is read-only!!")

"""
import xlwt
import csv
import os
import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
"""
#file = r"C:\Users\\tsezg523\Desktop\\test_file.xlsx"
"""xls = pd.ExcelFile(file)
header = tuple(xls.sheet_names)
data1 = pd.read_excel(xls, 'Sheet1')
data2 = pd.read_excel(xls, 'Sheet2')
# df = pd.DataFrame(data, columns=['Product'])
print("First output:", data1)

# Show the column Product only for the first row
print("Second output:", data1['Price'].iloc[0])

# Modify the data from the column Product for the first row
data1.iloc[0, data1.columns.get_loc('Price')] = 500
print("Third output:", data1)


print("Forth output:", data2)

# Print sheet names
print("Sheet names:", xls.sheet_names)

# Generate a dictionary of DataFrames
dfs = {sh: xls.parse(sh) for sh in xls.sheet_names}
print("new:", dfs.keys())
# Print all sheets
for df in dfs:
    print(dfs[df])
"""
# TODO: fix the ExcelWriter instantiate problem
# Try to write new data into the same excel file
# writer = pd.ExcelWriter(file, engine = 'xlsxwriter')
# Modify one of the data before writing back to the file
# data1.to_excel(writer, sheet_name = 'Sheet1', index = False, header = True)
# writer.save()
# writer.close()

# Test new method
